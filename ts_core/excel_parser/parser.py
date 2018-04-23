from openpyxl import *
import json
import xml.etree.ElementTree as ET
import xml.dom.minidom

SUMO_ATTR_COLUMN = 0    # Column A
SUMO_FILE_COLUMN = 1    # Column B
CATEGORY_COLUMN = 2
UNITS_COLUMN = 4        # Column E
VALUE_COLUMN = 5        # Column F for single entry data

MIN_ROW = 3  # Minimum row number where config information CAN be stored
MAX_ROW = 100    # Max row where data is stored TODO update?
MIN_COLUMN = 2

UNIT_CONVERT = {
    "time": {"Hours": 60*60, "Days": 60*60*24,
                "Months": 60*60*24*30.41,
                "Years": 60*60*24*30.41*12},  # How many seconds per time unit

    "angle": {"N": 90, "NE": 45, "E": 0, "SE": 315,
              "S": 270, "SW": 225, "W": 180, "NW": 135, },

}
# List of file extensions
ROUTE = "rou"
NET = "net"

SUMOFILE = 0
SUMOATTR = 1
USERVAL = 3  # First user_val
VALUNITS = 2

TAB_NAMES = ["Vehicle Type Customization", "General Settings", "Intersection Definition", "Branch Settings"]

ITYPE_BRANCHES = {
    "Cross": 4,
    "T": 3,
    "Y": 3,
}

OUTPUT_DICT = {
    "SUMOCFG": {},
    "Branches": {},
    "Intersections": {},
}


def parse_general(sheet):
    """
    Parses the General Settings Tab

    Fills out OUTPUT_DICT
    :param sheet: openpyxl sheet - The General Settings tab
    :return: None
    """
    # Get the Configuration name.
    input_config_name = get_input_from_row(sheet, MIN_ROW)     # Configuration Name
    # Write Config names to the output dict
    if input_config_name[SUMOFILE] == "SUMOCFG" and input_config_name[SUMOATTR] == "input":
        config_name = input_config_name[USERVAL]
        OUTPUT_DICT["SUMOCFG"]["input"] = {"net_file": "%s.net.xml" % config_name,
                                           "route_files": "%s.rou.xml" % config_name}
    else:
        print("Config name Error")  # TODO fill out

    # Simulation length
    input_runtime_length = get_input_from_row(sheet, MIN_ROW+1)

    if input_runtime_length[SUMOFILE] == "SUMOCFG" and input_runtime_length[SUMOATTR] == "time.end":
        sim_seconds = input_runtime_length[USERVAL] * UNIT_CONVERT["time"][
            input_runtime_length[VALUNITS]]  # TODO check if units is in time_convert?
        OUTPUT_DICT["SUMOCFG"]["time"] = {"begin": "0", "end": "%d" % sim_seconds}
    else:
        print("TIME ERROR", input_runtime_length)  # TODO fill out

    # Overall Traffic Demand
    # TODO Where does this go in the xmls if anywhere?
    # input_traffic_demand = get_input_from_row(sheet, MIN_ROW+2)
    return config_name


def parse_intersection(sheet):

    """
    
    Parameters
    ----------
    sheet: openpyxl sheet - The General Intersection Settings tab

    Returns
    -------
    String - The selected intersection type. Key in ITYPE_BRANCHES
    
    """

    # Intersection type (Cross, T, Y, etc)
    input_intersection_type = get_input_from_row(sheet, MIN_ROW)
    intersection_type = input_intersection_type[USERVAL]
    if intersection_type not in ITYPE_BRANCHES.keys():
        print("Unrecognized Intersection type: %s" % intersection_type)
        return None
    OUTPUT_DICT["Intersections"]["I0"] = {}

    # TODO Add in support for multiple intersections.
    # Initialize the intersection in The output Dict
    intersection_name = "I%d" % 0
    OUTPUT_DICT["Intersections"][intersection_name]["id"] = 0
    OUTPUT_DICT["Intersections"][intersection_name]["type"] = "traffic_light"

    return intersection_type


def parse_branches(sheet, intersection_type):
    """
    Parse the Branch Settings tab

    Fills out OUTPUT_DICT with all information needed to parse the branches of the intersection.
    :param sheet:
    :param intersection_type: The intersection Type (Cross, T, Y, etc. Needs to be a key in ITYPE_BRANCHES
    :return: None
    """
    # Get the number of branches
    num_branches = ITYPE_BRANCHES[intersection_type]

    # Initialize the output dictionary with the branches
    for branch_num in range(1, num_branches+1):
        b_name = "B%d" % branch_num # Branch name is Bn where n is the branches ID
        OUTPUT_DICT["Branches"][b_name] = {}

    for row in range(MIN_ROW, MAX_ROW):  # For each row that may contain data
        if row_has_data(sheet, row):        # Check if this row has data
            input_data = get_input_from_row(sheet, row, num_branches)   # Get the data
            units = input_data[VALUNITS].lower()


            # Fill out OUTPUT_DICT with data from each branch for this attribute
            for branch_num in range(0, num_branches):
                b_name = "B%d" % (branch_num+1)
                user_value = input_data[USERVAL+branch_num]
                if units in UNIT_CONVERT.keys():
                    user_value = UNIT_CONVERT[units][user_value]
                # TODO normalize the data using the units
                # TODO verify inbound lanes sum up to numLanes

                # write to output
                OUTPUT_DICT["Branches"][b_name][input_data[SUMOATTR]] = user_value


def row_has_data(sheet, row_num):
    """
    Checks whether or not this row contains a user-entered value
    :param sheet: openpyxl sheet - The sheet to use
    :param row_num: The row to analyze
    :return: Boolean - True if there is data, False otherwise
    """

    # Get the data from the row
    rows = sheet.iter_rows(min_col=1, min_row=row_num, max_col=15, max_row=row_num)
    for row in rows:    # FIXME Is there a better way to look at just 1 row
        # if there is no SUMO Attribute specified, then there is no user entered data available
        sumo_attr = row[SUMO_ATTR_COLUMN].value
        if sumo_attr is None:
            return False
        else:
            return True

    # should not be executed
    print("ERROR - row_has_data() did not find row")
    return False


def parse_stats(sheet):
    category_translate = {"Work Hours": "workHours", "City Gates": "cityGates",
                          "Bus Stations": "busStations", "Bus Lines": "busLines"}
    expected_entries = {"bracket": 3, "opening": 2, "closing": 2, "street": 3, "entrance": 4,
                        "school": 7, "busStation": 3}
    sub_tags = {"population": "bracket", "workHours": "opening", "streets": "street", "cityGates": "entrance", "schools": "school",
                "busStations": "busStation"}

    root = ET.Element("city")
    current_category = None
    category_root = None
    data = None
    num_entries = 15    # Maximum number of entries for any category TODO global

    for row in range(MIN_ROW, 100):
        # prettyprint(root)

        if row_has_category(sheet, row) is not None:
            # If a category has been parsed, put it in the xml
            if data is not None:
                for tag in data.keys():
                    attributes = data[tag]

                    if "TAG" in attributes.keys():
                        sub_tag = attributes["TAG"].lower()
                        del attributes["TAG"]
                    else:
                        sub_tag = tag.split("_")[0]

                    if sub_tag in expected_entries.keys() and len(attributes) < expected_entries[sub_tag]:
                        continue
                    ET.SubElement(category_root, sub_tag, attrib={attr: str(attributes[attr]) for attr in attributes})

                # prettyprint(root)
            current_category = row_has_category(sheet, row)
            if current_category in category_translate.keys():
                current_category = category_translate[current_category]
            else:
                current_category = current_category.lower().strip(" ")
            category_root = ET.SubElement(root, current_category)
            data = None
            continue
        elif category_root is None:
            continue

        if current_category in ["general", "parameters"]:
            if row_has_data(sheet, row):
                input_data = get_input_from_row(sheet, row, cols=1)
                category_root.attrib[input_data[SUMOATTR]] = str(input_data[USERVAL])

        elif current_category in ["population", "workHours", "streets", "cityGates", "schools", "busStations"]:
            sub_tag = sub_tags[current_category]

            if row_has_data(sheet, row):
                input_data = get_input_from_row(sheet, row, cols=100)
                if data is None:
                    data = {"%s_%d" % (sub_tag, i): {} for i in range(0, num_entries)}

                for i in range(0, len(input_data)-3):
                    attribute = input_data[SUMOATTR]
                    if attribute == "edge2":  # Deal with inbound/outbound attributes since sumo uses i and o notation
                        data["%s_%d" % (sub_tag, i)]["edge"] += {"Inbound": "i", "Outbound": "o"}[input_data[USERVAL+i]]
                    else:
                        data["%s_%d" % (sub_tag, i)][attribute] = str(input_data[USERVAL+i])

        elif current_category == "workHours":

            pass
    return root


def prettyprint(root):
    try:
        xml1 = xml.dom.minidom.parseString(ET.tostring(root, encoding='utf8', method='xml').decode())
        print(xml1.toprettyxml())
    except:
        pass


def row_has_category(sheet, row_num):
    """
    Checks to see if this row starts a new category
    :param sheet: openpyxl sheet - The sheet to use
    :param row_num: The row to analyze
    :return: Boolean - True if there is a new category, False otherwise
    """

    # Get the data from the row
    rows = sheet.iter_rows(min_col=1, min_row=row_num, max_col=15, max_row=row_num)
    for row in rows:    # FIXME Is there a better way to look at just 1 row
        # If there is text in the Category Column then it is a new category
        class_name = row[CATEGORY_COLUMN].value
        if class_name is None:
            return None
        else:
            return class_name

    # should not be executed
    print("ERROR - row_has_category() did not find any rows")
    return None


def get_input_from_row(sheet, row_num, cols=1):
    """
    Retrieves the input from the row and returns a tuple of all the relevant data
    :param sheet:
    :param row_num:
    :param cols:
    :return: Tupple containing the data -
              formatted as (sumo_file, sumo_attribute, units, user_val1, user_val2...)
    """
    output = []
    rows = sheet.iter_rows(min_col=1, min_row=row_num, max_col=VALUE_COLUMN+cols, max_row=row_num)

    for row in rows:
        # Determine the SUMO_FILE attribute
        sumo_file = row[SUMO_FILE_COLUMN].value
        if sumo_file is None: sumo_file = "N/A"
        output.append(sumo_file)

        # Determine the SUMO_ATTR attribute
        sumo_attr = row[SUMO_ATTR_COLUMN].value
        if sumo_attr is None: sumo_attr = "N/A"
        output.append(sumo_attr)

        # Determine the units
        units = row[UNITS_COLUMN].value  # TODO check if user entered units
        if units is None: units = "N/A"
        output.append(units)

        # Determine all user entered values
        for col in range(VALUE_COLUMN, VALUE_COLUMN+cols):
            user_value = row[col].value
            if user_value is None: break
            output.append(user_value)

        return tuple(output)


def run_parser(excel_file_path, outpath):
    """
    
    Parameters
    ----------
    excel_file_path
    outpath
    stats_file_path

    Returns
    -------

    """
    wb = load_workbook(filename=excel_file_path, data_only=True)

    config_name = parse_general(wb["General Settings"])


    type = parse_intersection(wb["Intersection Settings"])
    if parse_intersection is None:
        print("ERROR")
        return -1

    parse_branches(wb["Branch Settings"], type)


    # Context manager to dump parsed config to json
    with open(outpath+ "/" + config_name + "_parsed.json", 'w') as outfile:
        json.dump(OUTPUT_DICT, outfile, indent=4,)

    stats_root_node = parse_stats(wb["Advanced Customization"])


    stats_xml = xml.dom.minidom.parseString(ET.tostring(stats_root_node, encoding='utf8', method='xml').decode())

    return config_name, OUTPUT_DICT, stats_xml.toprettyxml()


if __name__ == "__main__":
    run_parser('Configuration_template.xlsx', './test_dir')

